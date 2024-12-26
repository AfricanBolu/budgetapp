import os
import subprocess
import xlsxwriter
from time import sleep
from menu import Menu
from openpyxl import load_workbook
from installer import install_library

# List of libraries to install
libraries = ["xlsxwriter", "openpyxl"]

for lib in libraries:
    install_library(lib)

class BudgetCalculator:
    def __init__(self):
        self.initial = 0
        self.percent = []
        self.final = []
        self.types = []


    def add_data(self):
        # error handling for user percentage input
        while True:
            try:
                n = int(input('Enter number of percentages: '))
                if n <= 0:
                    print('Error: Number of percentages must be greater than 0')
                    continue
                break
            except ValueError:
                print('Error: Number of percentages must be a number')

        # for val in range(n):
        #     self.types.append(input('Enter names for each percentage: '))
        #     self.percent.append(getPercentages(f'Enter percentage for {self.types[-1]}: '))

        for val in range(n):
            while True:
                type_name = input('Enter names for each percentage: ')
                if not type_name:
                    print('Error: Name cannot be empty')
                    continue
                if type_name in self.types:
                    print(f'Error: {type_name} already exists')
                    update = input('Do you want to update? (y/n): ')
                    if update.lower() == 'y':
                        idx = self.types.index(type_name)
                        new_percent = self.getPercentages(f'Enter new percentage for {type_name}: ')

                        if new_percent < len(self.percent):
                            self.percent[idx] = new_percent
                        else:
                            self.percent.append(new_percent)
                        break
                    else:
                        continue
                else:
                    self.types.append(type_name)
                    self.percent.append(self.getPercentages(f'Enter percentage for {type_name}: '))
                    break

        self.calculate()

    # def calculate(self):
    #     for val in range(len(self.percent)):
    #         self.final = self.initial * (self.percent[val] / 100)
    #     return self.final

    def calculate(self):
        print(f"\nBudget: {self.initial}")
        for i, category in enumerate(self.types):
            amount= self.initial * (self.percent[i] / 100)
            self.final.append(amount)
            print(f"{category}: ${amount:.2f} ({self.percent[i]}%)")

    def print(self):
        # print(f'Your amount entered: {self.total}')
        # print(f'Your percent entered: {self.percent}')
        for i, category in enumerate(self.types):
            print(f'Your final amount for {category} with {self.percent[i]}%: {self.final[i]:.2f}')

    def save(self, filename='budgets'):
        path =  os.path.join(os.path.expanduser('~'), 'Documents')
        work_path = os.path.join(path, f'{filename}.xlsx')
        # Get user profile name
        # user = os.getlogin()

        try:
            # Workbook() takes one required argument: file name
            if os.path.exists(work_path):
                option = input('File already exists. Do you want to overwrite? (y/n): ')
                if option.lower() != 'y':
                    rename = input('Enter new file name: ')
                    work_path = os.path.join(path, f'{rename}.xlsx')

            workbook = xlsxwriter.Workbook(work_path)
            worksheet = workbook.add_worksheet()

            # Datas to be written
            # Initial Amount(A1), Types(A3), Percent(B3), Final(C3)

            # format sheet
            bold = workbook.add_format({'bold': True})
            money = workbook.add_format({'num_format': '$#,##0.00'})
            percent_sign = workbook.add_format({'num_format': '.00%'})

            # Initial Amount(A1)
            worksheet.write('A1', 'Initial Amount', bold)
            worksheet.write('A2', self.initial, money)
            worksheet.write('B1', 'After Deductions', bold)
            worksheet.write('B2', self.getCurrent(), money)

            # Headers
            worksheet.write('A4', 'Types', bold)
            worksheet.write('B4', 'Percent%', bold)
            worksheet.write('C4', 'Final', bold)

            # Starting rows and columns
            row = 4
            col = 0

            # iterate over the types, percent and final
            for val in range(len(self.types)):
                worksheet.write_string(row, col, self.types[val])
                worksheet.write_number(row, col + 1, self.percent[val] / 100, percent_sign)
                worksheet.write_number(row, col + 2, self.final[val], money)
                row += 1

            # calc sum of final
            worksheet.write('A' + str(row + 1), 'Total', bold)
            worksheet.write('C' + str(row + 1), '=SUM(C5:C' + str(row) + ')', money)

            workbook.close()
            print(f'File saved to {work_path}')

            # ope file
            try:
                if os.name == 'nt':
                    os.startfile(work_path)
                elif os.name == 'posix':
                    subprocess.run(['open', work_path])

            except Exception as e:
                print(f'File could not be opened: {e}')
        except PermissionError:
            print('Error: Could not write to file. Check if file is open in another program.')
            # return None
        except Exception as e:
            print(f'Error: {e}')

    def load_file(self, filename='budgets'):
        path =  os.path.join(os.path.expanduser('~'), 'Documents')
        work_path = os.path.join(path, f'{filename}.xlsx')

        if os.path.exists(work_path):
            try:
                workbook = load_workbook(work_path)
                sheet = workbook.active

                # clear data
                self.types = []
                self.percent = []
                self.final = []

                # get data
                self.initial = float(sheet['A2'].value)

                for row in sheet.iter_rows(min_row=5, values_only=True):
                    if row[0] is not None and row[1] is not None:
                        self.types.append(row[0])
                        self.percent.append(float(row[1]))
                        self.final.append(self.initial * float(row[1]))

                print(f'File loaded from {work_path}\n')
                self.print()
                return True
            except Exception as e:
                print(f'Error loading file: {e}')
                return False
        else:
            print('File does not exist')
            return None

    def getPercentages(self,prompt):
        while True:
            try:
                percent = float(input(prompt))
                if percent < 0 or percent > 100:
                    print('Error: Percentage must be between 0 and 100')
                    continue
                return percent
            except ValueError:
                print('Error: Percentage must be a number')

    # def getCurrent(self):
    #     return self.initial - sum(self.final)
    def getCurrent(self):
        # Total calculated after deductions
        return sum(self.initial * (p / 100) for p in self.percent)

def run():
    # percent = [10, 40, 50]
    # percent = []
    # types = []
    # finals = []
    running = True



    app = BudgetCalculator()
    # app.add_data()
    # amount = app.initial
    # percent = app.percent
    # types = app.types
    # finals = app.final
    # app.calculate()
    # app.print()
    # app.save(filename)

    menu = Menu()

    while running:
        menu.display_menu()
        choice = menu.get_choice()
        match choice:
            case 1:
                while True:
                    try:
                        app.initial = float(input('Enter your initial amount: '))
                        if app.initial < 0:
                            print('Error: Amount cannot be negative')
                            continue
                        break
                    except ValueError:
                        print('Error: Amount must be a number')
                app.add_data()
                app.calculate()
                app.print()
            case 2:
                filename = input('Enter your filename: ')
                app.save(filename)
            case 3:
                filename = input('Enter file name: ')
                print(f"Attempting to load file: {filename}")
                result = app.load_file(filename)
                print(f"Load file result: {result}")  # Debugging
                if result:  # Check explicitly for success
                    user_input = input("Would you want to make changes? (y/n): ")
                    if user_input.lower() == 'y':
                        app.print()
                        app.add_data()
                        app.calculate()
                        app.print()
                        save_choice = input("Do you want to save the changes? (y/n): ")
                        if save_choice.lower() == 'y':
                            app.save(filename)
                else:
                    print("File could not be loaded.")
            case 4:
                running = False
            case _:
                print('Error: Invalid choice')



    # print('percentages entered: ', percent)
    # sleep(2)
    # os.system('cls' if os.name == 'nt' else 'clear')




if __name__ == '__main__':
    run()

# print(results)